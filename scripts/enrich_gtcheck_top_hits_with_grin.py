#!/usr/bin/env python3
"""
Annotate ranked gtcheck top hits with GRIN metadata.

This script is meant for the common post-processing step where a ranked
`gtcheck_top_hits.tsv` table is enriched with accession metadata from the USDA
GRIN Global BrAPI endpoint. It normalizes PI accession formatting, caches GRIN
lookups, and can write both TSV and XLSX outputs.

BrAPI endpoint usage adapted from get_displayNames_list.py
by Rex Nelson, Ames, IA Rex.Nelson@usda.gov
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import time
import urllib.parse
import urllib.request
from urllib.error import HTTPError
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    load_workbook = None


DEFAULT_API_BASE = "https://npgsweb.ars-grin.gov/gringlobal/brapi/v2/germplasm"
INSERT_COLUMNS = [
    "genotyped_sample",
    "PLANT NAME",
    "TAXONOMY",
    "ORIGIN",
    "GRIN ID",
    "GRIN LOOKUP STATUS",
    "GRIN MATCH METHOD",
]
PI_PATTERN = re.compile(r"^\s*PI\s*([0-9]+)\s*([A-Za-z]+)?\s*$", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add GRIN metadata to a gtcheck top-hits table.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("-i", "--input", required=True, help="Input TSV/CSV/XLSX from ranked gtcheck hits.")
    parser.add_argument(
        "--sample-column",
        default=None,
        help="Column containing the panel accession or genotype name. Defaults to auto-detect.",
    )
    parser.add_argument(
        "--crop",
        default="soybean",
        help="commonCropName used in the GRIN BrAPI query.",
    )
    parser.add_argument(
        "--api-base",
        default=DEFAULT_API_BASE,
        help="GRIN BrAPI germplasm endpoint.",
    )
    parser.add_argument(
        "--cache-json",
        default=None,
        help="Optional JSON cache path. Defaults to '<input stem>.grin_cache.json'.",
    )
    parser.add_argument(
        "--output-tsv",
        default=None,
        help="Output TSV path. Defaults to '<input stem>.grin_enriched.tsv'. Use 'none' to disable.",
    )
    parser.add_argument(
        "--output-xlsx",
        default=None,
        help="Output XLSX path. Defaults to '<input stem>.grin_enriched.xlsx'. Use 'none' to disable.",
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Do not query GRIN. Use only cached metadata already present in --cache-json.",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.0,
        help="Optional delay between uncached API requests.",
    )
    parser.add_argument(
        "--timeout-seconds",
        type=float,
        default=30.0,
        help="Timeout for each GRIN API request.",
    )
    return parser.parse_args()


def collapse_spaces(text: str) -> str:
    return " ".join(text.strip().split())


def normalize_sample_name(sample: str) -> str:
    clean = collapse_spaces(sample)
    match = PI_PATTERN.match(clean)
    if not match:
        return clean
    number = match.group(1)
    suffix = (match.group(2) or "").upper()
    if suffix:
        return f"PI {number} {suffix}"
    return f"PI {number}"


def build_accession_candidates(sample: str) -> list[str]:
    normalized = normalize_sample_name(sample)
    match = PI_PATTERN.match(normalized)
    candidates = [normalized]
    if match:
        number = match.group(1)
        suffix = (match.group(2) or "").upper()
        if suffix:
            candidates.append(f"PI {number}{suffix}")
    clean = collapse_spaces(sample)
    if clean not in candidates:
        candidates.append(clean)
    return list(dict.fromkeys(candidates))


def default_stem(path: Path) -> str:
    if path.suffix.lower() == ".gz":
        return path.with_suffix("").stem
    return path.stem


def resolve_outputs(args: argparse.Namespace) -> tuple[Optional[Path], Optional[Path], Path]:
    input_path = Path(args.input)
    stem = default_stem(input_path)
    parent = input_path.parent

    tsv = None if args.output_tsv == "none" else Path(args.output_tsv) if args.output_tsv else parent / f"{stem}.grin_enriched.tsv"
    xlsx = None if args.output_xlsx == "none" else Path(args.output_xlsx) if args.output_xlsx else parent / f"{stem}.grin_enriched.xlsx"
    cache = Path(args.cache_json) if args.cache_json else parent / f"{stem}.grin_cache.json"
    return tsv, xlsx, cache


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        if load_workbook is None:
            raise SystemExit("openpyxl is required to read .xlsx input files")
        workbook = load_workbook(path, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"No rows found in {path}")
        header = ["" if value is None else str(value) for value in rows[0]]
        out_rows: list[dict[str, str]] = []
        for row in rows[1:]:
            values = ["" if value is None else str(value) for value in row]
            values.extend([""] * max(0, len(header) - len(values)))
            out_rows.append(dict(zip(header, values[: len(header)])))
        return header, out_rows

    delimiter = "\t" if suffix in {".tsv", ".txt"} else ","
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        header = reader.fieldnames or []
        rows = [{key: value or "" for key, value in row.items()} for row in reader]
    if not header:
        raise ValueError(f"No header found in {path}")
    return header, rows


def detect_sample_column(header: list[str], requested: Optional[str]) -> str:
    if requested:
        if requested not in header:
            raise ValueError(f"--sample-column '{requested}' not found in input columns")
        return requested
    for name in ["genotyped_sample", "panel_sample", "top_panel_sample"]:
        if name in header:
            return name
    raise ValueError("Could not auto-detect sample column. Use --sample-column.")


def load_cache(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if isinstance(data, dict) and "entries" in data and isinstance(data["entries"], dict):
        return {cache_key(str(key)): dict(value) for key, value in data["entries"].items()}
    if isinstance(data, dict):
        return {cache_key(str(key)): dict(value) for key, value in data.items()}
    raise ValueError(f"Unexpected cache structure in {path}")


def save_cache(path: Path, cache: dict[str, dict[str, str]]) -> None:
    payload = {"entries": cache}
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, sort_keys=True)
        handle.write("\n")


def fetch_json(url: str, timeout_seconds: float) -> dict[str, object]:
    with urllib.request.urlopen(url, timeout=timeout_seconds) as response:
        return json.load(response)


def normalize_for_compare(text: str) -> str:
    return collapse_spaces(text).upper()


def cache_key(sample: str) -> str:
    return normalize_for_compare(normalize_sample_name(sample))


def choose_best_record(records: list[dict[str, object]], sample: str) -> Optional[dict[str, object]]:
    if not records:
        return None
    target = normalize_for_compare(normalize_sample_name(sample))

    def score(record: dict[str, object]) -> tuple[int, int]:
        accession = normalize_for_compare(str(record.get("accessionNumber", "")))
        display = normalize_for_compare(str(record.get("defaultDisplayName", "")))
        germplasm = normalize_for_compare(str(record.get("germplasmName", "")))
        exact_accession = int(accession == target)
        exact_name = int(display == target or germplasm == target)
        return (exact_accession, exact_name)

    return max(records, key=score)


def build_taxonomy(record: dict[str, object]) -> str:
    parts: list[str] = []
    genus = collapse_spaces(str(record.get("genus", "")))
    species = collapse_spaces(str(record.get("species", "")))
    species_authority = collapse_spaces(str(record.get("speciesAuthority", "")))
    subtaxa = collapse_spaces(str(record.get("subtaxa", "")))
    subtaxa_authority = collapse_spaces(str(record.get("subtaxaAuthority", "")))

    if genus:
        parts.append(genus)
    if species:
        parts.append(species)
    if species_authority:
        parts.append(species_authority)
    if subtaxa:
        parts.append(subtaxa)
    if subtaxa_authority:
        parts.append(subtaxa_authority)
    return " ".join(parts).strip()


def clean_origin(record: dict[str, object]) -> str:
    seed_source = collapse_spaces(str(record.get("seedSource", "")))
    if seed_source and seed_source.lower() != "none":
        seed_source = re.sub(r"^[A-Za-z ]*-\s*,\s*", "", seed_source)
        return seed_source.strip(" ,")
    code = collapse_spaces(str(record.get("countryOfOriginCode", "")))
    return code


def lookup_by_accession(candidate: str, crop: str, api_base: str, timeout_seconds: float) -> list[dict[str, object]]:
    params = urllib.parse.urlencode({"commonCropName": crop, "accessionNumber": candidate})
    try:
        payload = fetch_json(f"{api_base}?{params}", timeout_seconds)
    except HTTPError as exc:
        if exc.code in {404, 500}:
            return []
        raise
    return list(payload.get("result", {}).get("data", []))


def lookup_by_name(candidate: str, crop: str, api_base: str, timeout_seconds: float) -> list[dict[str, object]]:
    params = urllib.parse.urlencode({"commonCropName": crop, "germplasmName": candidate})
    try:
        payload = fetch_json(f"{api_base}?{params}", timeout_seconds)
    except HTTPError as exc:
        if exc.code in {404, 500}:
            return []
        raise
    return list(payload.get("result", {}).get("data", []))


def empty_annotation(normalized: str, status: str, method: str) -> dict[str, str]:
    return {
        "genotyped_sample": normalized,
        "PLANT NAME": "",
        "TAXONOMY": "",
        "ORIGIN": "",
        "GRIN ID": "",
        "GRIN LOOKUP STATUS": status,
        "GRIN MATCH METHOD": method,
    }


def record_to_annotation(record: dict[str, object], status: str, method: str, fallback_sample: str) -> dict[str, str]:
    accession = collapse_spaces(str(record.get("accessionNumber", ""))) or normalize_sample_name(fallback_sample)
    return {
        "genotyped_sample": accession,
        "PLANT NAME": collapse_spaces(str(record.get("defaultDisplayName", ""))) or collapse_spaces(str(record.get("germplasmName", ""))),
        "TAXONOMY": build_taxonomy(record),
        "ORIGIN": clean_origin(record),
        "GRIN ID": collapse_spaces(str(record.get("germplasmDbId", ""))),
        "GRIN LOOKUP STATUS": status,
        "GRIN MATCH METHOD": method,
    }


def lookup_annotation(
    sample: str,
    crop: str,
    api_base: str,
    timeout_seconds: float,
    cache: dict[str, dict[str, str]],
    offline: bool,
    sleep_seconds: float,
) -> dict[str, str]:
    normalized = normalize_sample_name(sample)
    key = cache_key(sample)
    cached = cache.get(key)
    if cached is not None:
        return dict(cached)
    if offline:
        return empty_annotation(normalized, "cache_miss_offline", "offline")

    accession_candidates = build_accession_candidates(sample)
    for candidate in accession_candidates:
        try:
            records = lookup_by_accession(candidate, crop, api_base, timeout_seconds)
        except Exception as exc:  # pragma: no cover - network errors depend on runtime
            cache[key] = empty_annotation(normalized, f"lookup_error:{type(exc).__name__}", "accession")
            return dict(cache[key])
        record = choose_best_record(records, sample)
        if record:
            annotation = record_to_annotation(record, "ok", "accession", sample)
            cache[key] = annotation
            if sleep_seconds > 0:
                time.sleep(sleep_seconds)
            return dict(annotation)
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)

    try:
        records = lookup_by_name(collapse_spaces(sample), crop, api_base, timeout_seconds)
    except Exception as exc:  # pragma: no cover - network errors depend on runtime
        cache[key] = empty_annotation(normalized, f"lookup_error:{type(exc).__name__}", "name")
        return dict(cache[key])

    record = choose_best_record(records, sample)
    if record:
        annotation = record_to_annotation(record, "ok", "name", sample)
    else:
        annotation = empty_annotation(normalized, "not_found", "name")
    cache[key] = annotation
    if sleep_seconds > 0:
        time.sleep(sleep_seconds)
    return dict(annotation)


def build_output_header(header: list[str], sample_column: str) -> list[str]:
    out_header: list[str] = []
    inserted = False
    for column in header:
        if column not in out_header:
            out_header.append(column)
        if column == sample_column:
            for extra in INSERT_COLUMNS:
                if extra not in out_header:
                    out_header.append(extra)
            inserted = True
    if not inserted:
        out_header.extend([column for column in INSERT_COLUMNS if column not in out_header])
    return out_header


def enrich_rows(
    rows: list[dict[str, str]],
    sample_column: str,
    crop: str,
    api_base: str,
    timeout_seconds: float,
    cache: dict[str, dict[str, str]],
    offline: bool,
    sleep_seconds: float,
) -> list[dict[str, str]]:
    enriched: list[dict[str, str]] = []
    for row in rows:
        sample = row.get(sample_column, "")
        annotation = lookup_annotation(
            sample=sample,
            crop=crop,
            api_base=api_base,
            timeout_seconds=timeout_seconds,
            cache=cache,
            offline=offline,
            sleep_seconds=sleep_seconds,
        )
        merged = dict(row)
        merged.update(annotation)
        enriched.append(merged)
    return enriched


def write_tsv(path: Path, header: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in header})


def write_xlsx(path: Path, header: list[str], rows: list[dict[str, str]]) -> None:
    if Workbook is None:
        raise SystemExit("openpyxl is required to write .xlsx output files")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "grin_enriched_gtcheck"
    worksheet.append(header)
    for row in rows:
        worksheet.append([row.get(column, "") for column in header])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(path)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    tsv_path, xlsx_path, cache_path = resolve_outputs(args)

    header, rows = read_table(input_path)
    sample_column = detect_sample_column(header, args.sample_column)
    cache = load_cache(cache_path)
    enriched_rows = enrich_rows(
        rows=rows,
        sample_column=sample_column,
        crop=args.crop,
        api_base=args.api_base,
        timeout_seconds=args.timeout_seconds,
        cache=cache,
        offline=args.offline,
        sleep_seconds=args.sleep_seconds,
    )
    out_header = build_output_header(header, sample_column)

    if tsv_path is not None:
        write_tsv(tsv_path, out_header, enriched_rows)
        print(f"Wrote: {tsv_path}")
    if xlsx_path is not None:
        write_xlsx(xlsx_path, out_header, enriched_rows)
        print(f"Wrote: {xlsx_path}")

    save_cache(cache_path, cache)
    print(f"Wrote: {cache_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
